from openpyxl import load_workbook
import pandas as pd

# Carregar Excel com valores (não fórmulas)
wb = load_workbook('base_cone.xlsx', data_only=True)
ws = wb['CONE']

# Estrutura: linha 0 = times, linha 1 = escopo, linha 2 = M/P headers
# Colunas: E=Semana, F=GOL_M, G=GOL_P, H=TERA_M, I=TERA_P, etc.

# Extrair header com times
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
times = []
for i, val in enumerate(header_row):
    if val and val not in [None, 'Semana', 'ESCOPO']:
        times.append(val)

print(f"Times encontrados: {times}")

# Mapeamento de colunas (baseado na estrutura observada)
# Coluna E (idx 4) = Semana
# Pares de colunas: M, P para cada time
col_mapping = {
    'GOL': {'M': 'F', 'P': 'G'},
    'TERA': {'M': 'H', 'P': 'I'},
    'OPTIMUS': {'M': 'J', 'P': 'K'},
    'NIVUS': {'M': 'L', 'P': 'M'},
    'TAOS': {'M': 'N', 'P': 'O'},
    'FUSCA': {'M': 'P', 'P': 'Q'},
    'UP': {'M': 'R', 'P': 'S'}
}

# Converter para índice de coluna (A=1, B=2, etc.)
def col_to_idx(col):
    return ord(col) - ord('A') + 1

# Extrair dados
data_rows = []
for row_idx in range(4, ws.max_row + 1):  # Começa na linha 4 (dados)
    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    semana = row[4]  # Coluna E
    
    if semana is None:
        continue
    
    for time in times:
        if time in col_mapping:
            m_col = col_to_idx(col_mapping[time]['M'])
            p_col = col_to_idx(col_mapping[time]['P'])
            
            melhor_caso = row[m_col - 1] if m_col - 1 < len(row) else None
            pior_caso = row[p_col - 1] if p_col - 1 < len(row) else None
            
            if melhor_caso is not None or pior_caso is not None:
                data_rows.append({
                    'Semana': semana,
                    'Time': time,
                    'Melhor caso': melhor_caso,
                    'Pior caso': pior_caso
                })

# Criar DataFrame
df = pd.DataFrame(data_rows)

# Salvar CSV
df.to_csv('base_cone_novo.csv', index=False, encoding='utf-8-sig')
print(f"\nCSV gerado com {len(df)} linhas")
print(df.head(20))
