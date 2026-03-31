from openpyxl import load_workbook
from datetime import datetime
import json

def format_date(dt):
    """Converte datetime para formato DD/MM/YYYY HH:MM"""
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt
    if isinstance(dt, datetime):
        return dt.strftime('%d/%m/%Y %H:%M')
    return str(dt)

# Carregar Excel com valores (não fórmulas)
wb = load_workbook('base_cone.xlsx', data_only=True)
ws = wb['BASE CONE']

# Ler dados - pulando header (linha 1)
data_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:  # Tipo de item vazio = fim dos dados
        break
    
    item = {
        "Type": row[0],           # Coluna A: Tipo de item
        "Key": str(row[1]),       # Coluna B: Chave da item
        "Summary": row[3],         # Coluna D: Resumo
        "Status": str(row[4]).upper() if row[4] else "UNKNOWN",  # Coluna E: Status (uppercase)
        "Team": str(row[5]) if row[5] else "",           # Coluna F: Time
        "Created": format_date(row[9]),                  # Coluna J: Criado
        "Resolved": format_date(row[12]) if row[12] else None,  # Coluna M: Resolvido
        "Release": str(row[13]) if row[13] else ""      # Coluna N: Release
    }
    data_rows.append(item)

print(f"Total itens extraídos: {len(data_rows)}")

# Verificar alguns exemplos
print("\nPrimeiros 3 itens:")
for item in data_rows[:3]:
    print(item)

# Salvar como JSON
with open('src/data.json', 'w', encoding='utf-8') as f:
    json.dump(data_rows, f, ensure_ascii=False, indent=2)

print(f"\nArquivo src/data.json atualizado com {len(data_rows)} itens")
